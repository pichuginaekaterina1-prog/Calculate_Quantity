from io import BytesIO

import pandas as pd
from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl.styles import PatternFill
from pydantic import BaseModel

MULTI_MARKER = "(Множественный выбор)_"
FALSE_LIKE_VALUES = {
    "",
    "0",
    "0.0",
    "false",
    "no",
    "none",
    "nan",
    "нет",
    "не выбрано",
    "не выбран",
}

app = FastAPI(title="Survey Cross-tab API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

storage = {
    "df": None,
    "columns": [],
    "questions_meta": [],
    "question_map": {},
}


class CrosstabRequest(BaseModel):
    row_questions: list[str] | None = None
    col_questions: list[str] | None = None
    metrics: list[str] | None = None
    significance: bool = False
    combine_columns: bool = False
    row_question: str | None = None
    col_question: str | None = None
    selected_row_values: list[str] | None = None
    selected_col_values: list[str] | None = None
    scale5: bool = False


@app.get("/")
def root():
    return {
        "message": "Survey Cross-tab API is running",
        "docs": "/docs",
        "health": "/health",
        "uploaded": storage["df"] is not None,
    }


@app.get("/health")
def health():
    return {
        "status": "ok",
        "uploaded": storage["df"] is not None,
    }


def normalize_value(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def is_answered_value(value) -> bool:
    return normalize_value(value) != ""


def is_selected_value(value) -> bool:
    normalized = normalize_value(value)
    return normalized != "" and normalized.lower() not in FALSE_LIKE_VALUES


def split_multi_column(column_name: str) -> tuple[str, str] | None:
    if MULTI_MARKER not in column_name:
        return None

    question_name, option_label = column_name.split(MULTI_MARKER, 1)
    question_name = question_name.strip()
    option_label = option_label.strip() or column_name.strip()
    return question_name, option_label


def ordered_unique_answers(series: pd.Series) -> list[str]:
    values = []
    seen = set()

    for value in series:
        normalized = normalize_value(value)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        values.append(normalized)

    return values


def build_question_catalog(df: pd.DataFrame) -> tuple[list[dict], dict[str, dict]]:
    questions = []
    question_map = {}

    for raw_column in df.columns.astype(str):
        multi_parts = split_multi_column(raw_column)

        if multi_parts:
            question_name, option_label = multi_parts
            question = question_map.get(question_name)
            if question is None:
                question = {
                    "name": question_name,
                    "kind": "multi",
                    "is_closed": True,
                    "columns": [],
                    "options": [],
                    "unique_count": 0,
                }
                questions.append(question)
                question_map[question_name] = question

            question["columns"].append(raw_column)
            question["options"].append({"label": option_label, "column": raw_column})
            question["unique_count"] = len(question["options"])
            continue

        series = df[raw_column]
        answers = ordered_unique_answers(series)
        non_blank_count = int(series.apply(is_answered_value).sum())
        unique_count = len(answers)
        unique_ratio = unique_count / non_blank_count if non_blank_count else 0

        question = {
            "name": raw_column,
            "kind": "single",
            "is_closed": non_blank_count > 0 and unique_count <= 20 and unique_ratio <= 0.5,
            "columns": [raw_column],
            "choices": answers,
            "unique_count": unique_count,
        }
        questions.append(question)
        question_map[raw_column] = question

    return questions, question_map


def get_question_definition(question_name: str) -> dict:
    question = storage["question_map"].get(question_name)
    if question is None:
        raise HTTPException(status_code=404, detail=f"Question not found: {question_name}")
    return question


def get_answered_mask(df: pd.DataFrame, question: dict) -> pd.Series:
    if question["kind"] == "single":
        return df[question["columns"][0]].apply(is_answered_value)

    mask = pd.Series(False, index=df.index)
    for option in question["options"]:
        mask = mask | df[option["column"]].apply(is_selected_value)
    return mask


def get_question_categories(df: pd.DataFrame, question: dict) -> list[dict]:
    if question["kind"] == "single":
        series = df[question["columns"][0]].apply(normalize_value)
        return [
            {"label": choice, "mask": series == choice}
            for choice in question.get("choices", [])
        ]

    return [
        {
            "label": option["label"],
            "mask": df[option["column"]].apply(is_selected_value),
        }
        for option in question["options"]
    ]


def build_standard_columns(
    df: pd.DataFrame,
    row_answered_mask: pd.Series,
    col_questions: list[dict],
) -> list[dict]:
    columns = [
        {
            "group": "Итого",
            "label": "Итого",
            "base": int(row_answered_mask.sum()),
            "mask": row_answered_mask,
        }
    ]

    for col_question in col_questions:
        for col_category in get_question_categories(df, col_question):
            base_mask = row_answered_mask & col_category["mask"]
            columns.append(
                {
                    "group": col_question["name"],
                    "label": col_category["label"],
                    "base": int(base_mask.sum()),
                    "mask": base_mask,
                }
            )

    return columns


def build_combined_columns(
    df: pd.DataFrame,
    row_answered_mask: pd.Series,
    col_questions: list[dict],
) -> list[dict]:
    columns = [
        {
            "group": "Итого",
            "label": "Итого",
            "base": int(row_answered_mask.sum()),
            "mask": row_answered_mask,
        }
    ]

    category_groups = [get_question_categories(df, col_question) for col_question in col_questions]
    if not category_groups:
        return columns

    combinations = [([], row_answered_mask)]
    for category_group in category_groups:
        next_combinations = []
        for labels, mask in combinations:
            for category in category_group:
                next_combinations.append((labels + [category["label"]], mask & category["mask"]))
        combinations = next_combinations

    for labels, mask in combinations:
        flat_label = " | ".join(labels)
        if len(labels) >= 2:
            group = labels[0]
            label = " | ".join(labels[1:])
        else:
            group = col_questions[0]["name"]
            label = flat_label

        base_mask = row_answered_mask & mask
        columns.append(
            {
                "group": group,
                "label": label,
                "flat_label": flat_label,
                "base": int(base_mask.sum()),
                "mask": base_mask,
            }
        )

    return columns


def is_significant_difference(count_a: int, base_a: int, count_b: int, base_b: int) -> bool:
    if base_a < 50 or base_b < 50 or base_a == 0 or base_b == 0:
        return False

    p1 = count_a / base_a
    p2 = count_b / base_b
    pooled = (count_a + count_b) / (base_a + base_b)
    variance = pooled * (1 - pooled) * ((1 / base_a) + (1 / base_b))
    if variance <= 0:
        return False

    z_score = abs((p1 - p2) / (variance ** 0.5))
    return z_score >= 1.959963984540054


def build_significance_styles(
    row_counts: dict[str, dict[str, int]],
    columns_by_key: dict[str, dict],
) -> dict[str, dict[str, str]]:
    styles = {}
    grouped_keys = {}

    for column_key, column in columns_by_key.items():
        if column["group"] == "Итого":
            continue
        grouped_keys.setdefault(column["group"], []).append(column_key)

    for row_key, counts in row_counts.items():
        row_styles = {}

        for group_keys in grouped_keys.values():
            for column_key in group_keys:
                base = columns_by_key[column_key]["base"]
                if base < 50:
                    row_styles[column_key] = "low_base"
                    continue

                higher = False
                lower = False
                current_count = counts[column_key]

                for other_key in group_keys:
                    if other_key == column_key:
                        continue

                    other_base = columns_by_key[other_key]["base"]
                    if other_base < 50:
                        continue

                    other_count = counts[other_key]
                    if not is_significant_difference(current_count, base, other_count, other_base):
                        continue

                    current_share = current_count / base
                    other_share = other_count / other_base
                    if current_share > other_share:
                        higher = True
                    elif current_share < other_share:
                        lower = True

                if higher and not lower:
                    row_styles[column_key] = "higher"
                elif lower and not higher:
                    row_styles[column_key] = "lower"

        styles[row_key] = row_styles

    return styles


def format_metric_value(metric: str, count: int, base: int) -> float | int:
    if metric == "count":
        return int(count)
    return int(round((count / base * 100) if base else 0))


def build_metric_table(
    df: pd.DataFrame,
    row_question: dict,
    col_questions: list[dict],
    metric: str,
    combine_columns: bool = False,
    significance: bool = False,
) -> dict:
    row_answered_mask = get_answered_mask(df, row_question)
    row_categories = get_question_categories(df, row_question)
    columns = (
        build_combined_columns(df, row_answered_mask, col_questions)
        if combine_columns
        else build_standard_columns(df, row_answered_mask, col_questions)
    )
    column_keys = [f"col_{index}" for index in range(len(columns))]
    columns_by_key = {key: column for key, column in zip(column_keys, columns)}

    rows = []
    row_counts = {}
    for row_category in row_categories:
        row_values = {}
        raw_counts = {}

        for column_key, column in columns_by_key.items():
            count = int((column["mask"] & row_category["mask"]).sum())
            raw_counts[column_key] = count
            row_values[column_key] = format_metric_value(metric, count, column["base"])
        rows.append({"label": row_category["label"], "values": row_values})
        row_counts[row_category["label"]] = raw_counts

    column_labels = {key: column["label"] for key, column in zip(column_keys, columns)}
    column_groups = {key: column["group"] for key, column in zip(column_keys, columns)}
    bases = {key: column["base"] for key, column in zip(column_keys, columns)}

    significance_styles = (
        build_significance_styles(row_counts, columns_by_key)
        if significance and metric == "percent"
        else {}
    )

    normalized_rows = []
    for row in rows:
        normalized_rows.append(
            {
                "label": row["label"],
                "values": row["values"],
                "styles": significance_styles.get(row["label"], {}),
            }
        )

    return {
        "row_question": row_question["name"],
        "col_question": ", ".join(col_question["name"] for col_question in col_questions),
        "row_kind": row_question["kind"],
        "col_kind": "combined",
        "metric": metric,
        "metric_label": "Количество (N)" if metric == "count" else "% по столбцу",
        "column_order": column_keys,
        "column_labels": column_labels,
        "column_groups": column_groups,
        "rows": normalized_rows,
        "bases": bases,
        "significance_enabled": significance and metric == "percent",
        "combine_columns": combine_columns,
        "base_description": "База - фактическое количество ответивших на вопрос в каждой группе.",
    }


def resolve_crosstab_request(req: CrosstabRequest) -> tuple[pd.DataFrame, list[dict], list[dict], list[str], bool, bool]:
    row_question_names = req.row_questions or ([req.row_question] if req.row_question else [])
    col_question_names = req.col_questions or ([req.col_question] if req.col_question else [])
    metrics = req.metrics or ["count", "percent"]
    metrics = [metric for metric in metrics if metric in {"count", "percent"}]

    if not row_question_names or not col_question_names:
        raise HTTPException(
            status_code=400,
            detail="At least one row question and one column question must be selected",
        )

    if not metrics:
        raise HTTPException(status_code=400, detail="At least one metric must be selected")

    df = storage["df"].copy()
    row_questions = [get_question_definition(question_name) for question_name in row_question_names]
    col_questions = [get_question_definition(question_name) for question_name in col_question_names]

    return df, row_questions, col_questions, metrics, req.combine_columns, req.significance


def build_tables_for_request(req: CrosstabRequest) -> dict:
    df, row_questions, col_questions, metrics, combine_columns, significance = resolve_crosstab_request(req)

    tables = [
        build_metric_table(df, row_question, col_questions, metric, combine_columns, significance)
        for row_question in row_questions
        for metric in metrics
    ]

    return {
        "tables": tables,
        "row_questions": [question["name"] for question in row_questions],
        "col_questions": [question["name"] for question in col_questions],
        "metrics": metrics,
        "combine_columns": combine_columns,
        "significance": significance,
    }


def append_table_to_sheet_rows(sheet_rows: list[list], table: dict):
    column_order = table["column_order"]
    sheet_rows.append([table["row_question"]])
    sheet_rows.append([table["metric_label"]])
    sheet_rows.append(["Срез", *[table["column_groups"][column_key] for column_key in column_order]])
    sheet_rows.append(["Варианты ответа", *[table["column_labels"][column_key] for column_key in column_order]])

    for row in table["rows"]:
        sheet_rows.append(
            [row["label"], *[row["values"].get(column_key, 0) for column_key in column_order]]
        )

    sheet_rows.append(["База", *[table["bases"].get(column_key, 0) for column_key in column_order]])
    sheet_rows.append([table["base_description"]])
    sheet_rows.append([])


def build_export_workbook(tables: list[dict]) -> BytesIO:
    output = BytesIO()
    metric_to_sheet = {
        "count": "Количество (N)",
        "percent": "% по столбцу",
    }
    fills = {
        "higher": PatternFill(fill_type="solid", fgColor="A4EC9E"),
        "lower": PatternFill(fill_type="solid", fgColor="F0B3B2"),
        "low_base": PatternFill(fill_type="solid", fgColor="DADAD8"),
    }

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for metric, sheet_name in metric_to_sheet.items():
            metric_tables = [table for table in tables if table["metric"] == metric]
            if not metric_tables:
                continue

            sheet_rows = []
            for table in metric_tables:
                append_table_to_sheet_rows(sheet_rows, table)

            pd.DataFrame(sheet_rows).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False,
            )

            worksheet = writer.sheets[sheet_name]
            row_idx = 1
            for table in metric_tables:
                data_start_row = row_idx + 3
                data_row_count = len(table["rows"])
                start_col = 2
                end_col = len(table["column_order"]) + 1

                if metric == "percent":
                    for excel_row in range(data_start_row, data_start_row + data_row_count):
                        row_offset = excel_row - data_start_row
                        for excel_col in range(start_col, end_col + 1):
                            column_offset = excel_col - start_col
                            cell = worksheet.cell(row=excel_row, column=excel_col)
                            cell.value = (cell.value or 0) / 100
                            cell.number_format = "0%"

                            style_key = table["rows"][row_offset].get("styles", {}).get(
                                table["column_order"][column_offset]
                            )
                            if style_key in fills:
                                cell.fill = fills[style_key]

                row_idx = data_start_row + data_row_count + 3

    output.seek(0)
    return output


@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    if file.content_type not in [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "application/octet-stream",
    ]:
        raise HTTPException(status_code=400, detail="Unsupported file type")

    content = await file.read()
    try:
        df = pd.read_excel(BytesIO(content))
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel: {exc}")

    if df.empty:
        raise HTTPException(status_code=400, detail="Empty Excel file")

    questions_meta, question_map = build_question_catalog(df)

    storage["df"] = df
    storage["columns"] = df.columns.astype(str).tolist()
    storage["questions_meta"] = questions_meta
    storage["question_map"] = question_map

    return {
        "rows": len(df),
        "columns": storage["columns"],
        "questions": [question["name"] for question in questions_meta],
    }


@app.get("/questions")
def questions():
    if storage["df"] is None:
        raise HTTPException(status_code=404, detail="No data uploaded")

    return {
        "questions": [question["name"] for question in storage["questions_meta"]],
        "items": storage["questions_meta"],
    }


@app.get("/choices")
def choices(question: str):
    if storage["df"] is None:
        raise HTTPException(status_code=404, detail="No data uploaded")

    question_def = get_question_definition(question)

    if question_def["kind"] == "single":
        return {"question": question, "choices": question_def.get("choices", [])}

    return {
        "question": question,
        "choices": [option["label"] for option in question_def.get("options", [])],
    }


@app.post("/crosstab")
def crosstab(req: CrosstabRequest):
    if storage["df"] is None:
        raise HTTPException(status_code=404, detail="No data uploaded")

    return build_tables_for_request(req)


@app.post("/export/xlsx")
def export_xlsx(req: CrosstabRequest):
    if storage["df"] is None:
        raise HTTPException(status_code=404, detail="No data uploaded")

    result = build_tables_for_request(req)
    workbook = build_export_workbook(result["tables"])

    return StreamingResponse(
        workbook,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="crosstab_export.xlsx"',
        },
    )
